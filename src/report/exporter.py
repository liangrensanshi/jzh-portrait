"""Excel export for household metrics, labels, and audit results."""
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from ..io.database import get_conn
from ..compute.metrics import _period_label

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output"

QUALITY_FILL = {
    "A": PatternFill("solid", fgColor="C8E6C9"),
    "B": PatternFill("solid", fgColor="F0F4C3"),
    "C": PatternFill("solid", fgColor="FFE0B2"),
    "D": PatternFill("solid", fgColor="FFCDD2"),
}


def export_summary(months: list, path: Optional[str] = None) -> str:
    period_start, period_end = _period_label(months)

    conn = get_conn()
    metrics = pd.read_sql(
        "SELECT * FROM household_metrics WHERE period_start=? AND period_end=?",
        conn, params=(period_start, period_end),
    )
    labels = pd.read_sql(
        "SELECT * FROM household_labels WHERE period_start=? AND period_end=?",
        conn, params=(period_start, period_end),
    )
    audit_counts = pd.read_sql(
        "SELECT hhid, COUNT(*) as rule_count FROM audit_results "
        "WHERE period_start=? AND period_end=? GROUP BY hhid",
        conn, params=(period_start, period_end),
    )
    households = pd.read_sql("SELECT hhid, hname, coun, vcode, urban_rural FROM raw_households", conn)
    villages = pd.read_sql("SELECT vcode, coun, counname, townname, vname FROM raw_villages", conn)
    conn.close()

    merged = households.merge(
        villages, on=["vcode", "coun"], how="left"
    ).merge(
        labels[["hhid", "label_income", "label_lifecycle", "label_mobility",
                "label_quality", "quality_score"]],
        on="hhid", how="left",
    ).merge(
        metrics[["hhid", "income_total", "consume_total", "consume_food",
                 "family_size_permanent", "ledger_count", "max_gap_days"]],
        on="hhid", how="left",
    ).merge(
        audit_counts, on="hhid", how="left",
    )

    merged["rule_count"] = merged["rule_count"].fillna(0).astype(int)

    if path is None:
        OUTPUT_DIR.mkdir(exist_ok=True)
        path = str(OUTPUT_DIR / f"summary_{period_start}_{period_end}.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="汇总", index=False)
        ws = writer.sheets["汇总"]

        for col_idx in range(1, len(merged.columns) + 1):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D0E4F7")
            cell.alignment = Alignment(horizontal="center")

        if "label_quality" in merged.columns:
            q_col_idx = list(merged.columns).index("label_quality") + 1
            for row_idx, row in enumerate(merged.itertuples(), start=2):
                q = getattr(row, "label_quality", None)
                fill = QUALITY_FILL.get(q)
                if fill:
                    for col_idx in range(1, len(merged.columns) + 1):
                        ws.cell(row_idx, col_idx).fill = fill

        for col_idx, col_name in enumerate(merged.columns, start=1):
            max_len = max(
                len(str(col_name)),
                *[len(str(v)) for v in merged[col_name].fillna("").tolist()[:50]],
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    return path
